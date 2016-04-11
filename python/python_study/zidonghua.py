# _*_ coding: UTF-8 _*_
from xml.dom import minidom
# from importlib import reload
import imp
import xlrd
import openpyxl
import json
import requests
import sys
import HTMLParser
import os
import re
import codecs
import time
import datetime


reload(sys)
sys.setdefaultencoding('utf-8')
class OptionExcelData(object):
	"""对Excel进行操作，包括读取请求参数，和操作结果"""
	def __init__(self, excelFile, excelPath=''):
		self.excelFile = excelFile
		self.excelPath = excelPath
		self.caseList = []


	def getCaseList(self, excelFile, excelPath=''):
		readExcel = xlrd.open_workbook(fileName)
		try:
			table = readExcel.sheet_by_index(0)
			trows = table.nrows
			for n in range(1,trows):
				tmpdict = {}
				tmpdict['id'] = n
				tmpdict['CityName'] = table.cell(n, 2).value
				tmpdict['CountryName'] = table.cell(n, 3).value
				tmpdict['Rspect'] = table.cell(n, 4).value
				self.caseList.append(tmpdict)
		except Exception as e:
			raise
		finally:
				pass
		return self.caseList


	def writeResult(self, resultBody, isSuccess, respTime, excelFile, theRow, theCol=5):
		writeExcel = openpyxl.load_workbook(excelFile)
		try:
			wtable = writeExcel.get_sheet_by_name('Sheet1')
			wtable.cell(row=theRow+1, column=theCol+1).value = resultBody
			wtable.cell(row=theRow+1, column=theCol+2).value = isSuccess
			wtable.cell(row=theRow+1, column=theCol+3).value = respTime
			writeExcel.save(excelFile)
		except Exception as e:
			raise
		finally:
			pass


class GetWeather(object):
	def __init__(self, serviceUrl, requestBody, headers):
		self.serviceUrl = serviceUrl
		self.requestBody = requestBody
		self.headers = headers
		self.requestResult = {}

	def getWeath(self, serviceUrl, requestBody, headers):
		timebefore = tiem.time()
		tmp = requests.post(serviceUrl, data=requestBody, headers=headers)
		timeend = time.time()
		tmptext = tmp.text
		self.requestResult['text'] = tmptext
		self.requestResult['time'] = round(timeend - timebefore, 2)
		return self.requestResult

class XmlReader:
	def __init__(self, testFile, testFilePath=''):
		self.fromXml = testFile
		self.xmlFilePath = testFilePath
		self.resultList = []

	def writeXmlData(self, resultBody, testFile, testFilePath=''):
		tmpXmlFile = codecs.open(testFile, 'w', 'utf-16')
		tmpLogFile = codecs.open(testFile+'.log', 'w', 'utf-16')

		tmp1 = re.compile(r'\<.*?\>')
		tmp2 = tmp1.sub('', resultBody['text'])
		html_parser = HTMLParser.HTMLParser()
		xmlText = html_parser.unescape(tmp2)

		try:
			tmpXmlFile.writelines(xmlText.strip())
			tmpLogFile.writelines('time: '+ str(resultBody['time']) + '\r\n')
			tmpLogFile.writelines('text: ' + resultBody['text'].strip())
		except Exception as e:
			raise
		finally:
			tmpXmlFile.close()
			tmpLogFile.close()

	def readXmlData(self, testFile, testFilePath=''):
		tmpXmlFile = minidom.parse(testFile)
		root = tmpXmlFile.documentElement
		tmpValue = root.getElementsByTagName('Status')[0].childNodes[0].data
		return tmpValue

if __name__ == '__main__':
	requsturl = 'http://www.webservicex.net/globalwether.asmx/GetWeather'
	requestHeaders = {"content-Type":"application/x-wwww-form-urlencoded"}
	fileName = u'test.xlsx'

	print fileName
	ed = OptionExcelData(fileName)
	testCaseList = ed.getCaseList(ed.excelFile)

	for caseDict in testCaseList:
		caseId = caseDict['id']
		CityName = caseDict['CityName']
		CountryName = caseDict['CountryName']
		rspect = caseDict['Rspect']
		requestBody = 'CityName=' + CityName + '&CountryName=' + CountryName
		print requestBody
		getWeather = GetWeather(requsturl, requestBody, requestHeaders)
		tmpString = getWeather.getWeath(getWeather.serviceUrl, getWeather.requestBody,getWeather.headers)

		xd = XmlReader(str(caseId) + '.xml')
		xd.writeXmlData(tmpString, xd.fromXml)
		response = xd.readXmlData(str(caseId) + '.xml')
		respTime = tmpString['time']
		if response == rspect:
			theResult = 'Pass'
		else:
			theResult = 'False'

		ed.writeCaseResult(response, theResult, respTime, fileName, caseId)